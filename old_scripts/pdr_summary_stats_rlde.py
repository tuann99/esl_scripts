###########################################################################################################################################
###########################################################################################################################################
# Michigan State University - Exposure Science Lab
# Creator: Tuan Nguyen
# Date Created: 2023-10-04
#
# Description:
#  - This is a script for extracting data from Thermo Scientific Personal DataRAM (pDR) text files, calculating summary statistics, and 
#    then creating an excel file with the data, summary stats, and chart.
#
# Requirements:
#  - Python 3.x
#  - Required Python packages: pandas, openpyxl, regex
#  - Input data files must be in the "new_files" directory within the specified input directory.
#  - Input data files should have "Subject #" in their file names.
#  - pDR serial numbers and corresponding names must be specified in the PDR_NUM_DICT dictionary within the script.
#
# Usage:
#  - If the data has been pulled off the pDR already, proceed to step 6, otherwise continue.
#  1. Plug the pDR into the computer and open the pDR Port software.
#  2. Click the "Data text" tab, then click the import icon (pdr with arrow pointing down).
#  3. Select the correct files to import and then press x
#  4. Ensure the folder you're importing to is "S:\\ExposureScienceLab\\Lead Dust\\Data\\pDR\\new_files"
#  5. Save the data with the correct subject ID in the file name. (Ex. Subject 1 Initial Aug 21-22 2023)
#  6. On windows, open the command prompt and type:
#     "python S:\ExposureScienceLab\scripts\pdr_summary_stats_rlde.py -i path\to\input_directory -o path\to\output_directory"
#     where "path\to\input_directory" is the path to the directory containing the pDR txt files and "path\to\output_directory" is the 
#     path to where you want the output files to be saved.
#
# Notes:
#  - Data will be exported to the respective subject folder in the pDR folder.
#  - If output directory is not specified, then by default the program will output to: "S:\ExposureScienceLab\Lead Dust\Data\pDR"
#  - If input directory is not specified, then by default the program will look for files in: 
#    "S:\\ExposureScienceLab\\Lead Dust\\Data\\pDR\\new_files"
#  - The script will create folders in the output directory specified for each subject ID.
#  - Typical usage will have the input directory be the new_files folder ("S:\\ExposureScienceLab\\Lead Dust\\Data\\pDR\\new_files") and 
#    the output directory be the pDR folder ("S:\\ExposureScienceLab\\Lead Dust\\Data\\pDR").
###########################################################################################################################################
###########################################################################################################################################

#!/usr/bin/env python3

import pandas as pd
import os
import string
import argparse
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference 
import regex as re
import shutil

# variables
def find_directory(drives, target):
    for drive in drives:
        directory = os.path.join(drive, target)
        if os.path.exists(directory):
            return directory
    return None

drives = ['%s:\\' % d for d in string.ascii_uppercase if os.path.exists('%s:\\' % d)]
targets = ['CHM\\ExposureScienceLab\\Lead Dust (rlde,hud)', 'ExposureScienceLab\\Lead Dust (rlde,hud)', 'Lead Dust (rlde,hud)']

for target in targets:
    directory = find_directory(drives, target)
    if directory is None:
        print(f"Directory '{target}' not found.")
    else:
        print(f"Directory found: {directory}")

DEFAULT_INPUT_DIRECTORY = os.path.join(directory, "Data\\pDR\\new_files")
DEFAULT_OUTPUT_DIRECTORY = os.path.join(directory, "Data\\pDR")
CUSTOM_HEADERS = ["record", "ug/m3", "Temp", "RHumidity", "AtmoPressure", "Flags", "time", "date"]
PDR_NUM_DICT = {
    "0115250158": "pdr_1",
    "0115249628": "pdr_2",
    "0115249629": "pdr_3",
    "0115250156": "pdr_4",
    # pdr 5
    "CM19342019": "pdr_6",
    # pdr 7
    "CM21092015": "pdr_8",
    "CM21102014": "pdr_9",
}

# start of script
def pdr_summary_stats_rlde(input_directory, subject_folder):
    for file in os.listdir(input_directory):
        if file.endswith(".txt"):
            
            # first create some file paths and extract some info
            file_path = os.path.join(input_directory, file) # create file path
            print(f"File found at: {file_path}.")
            base_name = os.path.splitext(os.path.basename(file))[0] # extracting name for xlsx file
            print(f"File name is: {base_name}.")
            subject_match = re.search(r"Subject \d+", base_name) # extracting subject ID. should be in format "Subject #"
            
            if subject_match: # check to make sure we got a subject id from the file name
                subject_id = subject_match.group(0)
                print(f"Subject ID found: {subject_id}.")
            else:
                print(f"Unable to find subject ID for {file_path}.")
                print(f"File was read as \"{base_name}.\"")
                print("Please ensure that the file name contains the subject ID as 'Subject #', where # is the subject number.")
                break

            # create a folder for the subject ID if it doesn't exist
            output_folder = os.path.join(subject_folder, subject_id)
            os.makedirs(output_folder, exist_ok=True)

            # copy file to correct subject folder
            copy = shutil.copy2(file_path, output_folder)
            setwd = os.chdir(output_folder)

            # create output file paths
            output_txt_path = os.path.join(output_folder, file)
            output_xlsx_path = os.path.join(output_folder, f"{base_name}.xlsx")

            # working w the xlsx file
            workbook = Workbook()

            # open the file and see if there is a serial number, if so, then go ahead with data extraction
            with open(output_txt_path, 'r') as z:
                lines = z.readlines()
                for line in lines: # parse file for serial number by looking for string after "Serial no.  "
                    if "Serial no." in line: # if serial number is found, then go ahead with data extraction
                        serial_num = line.split("Serial no.  \", ")[1].strip().replace('"', '')
                        print(f"{subject_id} was using: " + PDR_NUM_DICT[serial_num])
                        tmp = pd.read_csv(file_path, sep=",", skiprows=24, header=None, names=CUSTOM_HEADERS)
                        tmp["pdr name"] = PDR_NUM_DICT[serial_num]  # add column for pdr name
                        tmp["time"] = tmp["time"].str.strip()  # Remove leading and trailing whitespace
                        tmp["date"] = tmp["date"].str.strip()
                        summary_stats = tmp["ug/m3"].describe() # summary statistics

                        # write thedata to xlsx file in 'Data' sheet
                        data_sheet = workbook.create_sheet(title='Data') # create sheet for data
                        data_sheet.append(CUSTOM_HEADERS) # add headers to data sheet
                        for idx, row in tmp.iterrows():
                            data_sheet.append(row.tolist())

                        # write the summary statistics to xlsx file in a new sheet
                        summary_sheet = workbook.create_sheet(title='Summary Statistics') # create sheet for summary statistics
                        summary_stats_list = summary_stats.reset_index().values.tolist()
                        for row in summary_stats_list:
                            summary_sheet.append(row)
                        
                        # create the chart on summary sheet
                        chart = LineChart()
                        chart.title = "ug/m3 Over Time"
                        chart.y_axis.title = "ug/m3"
                        chart.x_axis.title = "Time"
                        y = Reference(data_sheet, min_col=2, min_row=2, max_col=2, max_row=len(tmp) + 1)
                        x = Reference(data_sheet, min_col=7, min_row=2, max_row=len(tmp) + 1)
                        chart.add_data(y, titles_from_data=True)
                        chart.set_categories(x)
                        summary_sheet.add_chart(chart, "A10")

                        # remove the default sheet
                        default_sheet = workbook["Sheet"]
                        workbook.remove(default_sheet)

                        workbook.save(output_xlsx_path)
                        print(f"Data and Summary statistics saved to {output_xlsx_path}")
        
        # check if output exists and has stuff in it, if so, then delete the file from the new files folder
        if file in os.listdir(output_folder):
            if os.path.getsize(output_xlsx_path) > 0:
                os.remove(file_path)
                print(f"File \"{file}\" deleted from \"{input_directory}.\"")
            else:
                print(f"File {file} is empty. Please check the file and try again.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="RLDE-specific extraction of data from pDR txt files, calculate summary statistics, and create excel file with data, summary stats, and chart.")
    parser.add_argument("-i", "--input_directory", type=str, default=DEFAULT_INPUT_DIRECTORY, help="The directory containing the pDR txt files.")
    parser.add_argument("-o", "--output_directory", type=str, default=DEFAULT_OUTPUT_DIRECTORY, help="The directory to save the output files to.")
    args = parser.parse_args()
    pdr_summary_stats_rlde(args.input_directory, args.output_directory)