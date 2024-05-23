#!/usr/bin/env python3

###########################################################################################################################################
###########################################################################################################################################
# Michigan State University - Exposure Science Lab
# Creator: Tuan Nguyen
# Date Created: 2023-10-04
#
# Description:
#  - This is a script for extracting data from Thermo Scientific Personal DataRAM (pDR) text files, calculating summary statistics, and 
#    then creating an excel file with the data, summary stats, and chart.
#
# Requirements:
#  - Python 3.x
#  - Required Python packages: pandas, openpyxl, regex
#  - If using the script for the RLDE project, input data files must be in the "../Lead Dust/Data/pDR/new_files" directory.
#  - Input data files should have a subject # in their file names formatted as 'Subject #' (Example: Subject 3).
#  - pDR serial numbers and corresponding names must be specified in the PDR_NUM_DICT dictionary within the script.
#
# Usage:
#  - "python path\to\pdr_summary_stats.py"
#  - There are optional flags for specifying an input directory ("-i path\to\input_directory"),
#    specifying an output directory ("-o path\to\output_directory"),
#    and specifying that the analysis will be for the RLDE project (-r). 
#
# Notes:
#  - Data will be exported to the respective subject folder in the pDR folder.
#  - If output directory is not specified, then by default the program will output to: "S:\ExposureScienceLab\Lead Dust\Data\pDR"
#  - If input directory is not specified, then by default the program will look for files in: 
#    "S:\\ExposureScienceLab\\Lead Dust\\Data\\pDR\\new_files"
#  - The script will create folders in the output directory specified for each subject ID.
#  - Typical usage will have the input directory be the new_files folder ("S:\\ExposureScienceLab\\Lead Dust\\Data\\pDR\\new_files") and 
#    the output directory be the pDR folder ("S:\\ExposureScienceLab\\Lead Dust\\Data\\pDR").
###########################################################################################################################################
###########################################################################################################################################

# Need to be installed
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference 
import pandas as pd
import regex as re

# Built into Python
import os
import string
import argparse
import shutil
import pathlib

def append_rows_to_sheet(sheet, df):
    '''
    Description:
        - This function appends the rows from a dataframe to a sheet.
    Parameters:
        - sheet: the sheet to append the rows to
        - df: the dataframe containing the rows to append
    Returns:
        - None
    '''
    for row in df.itertuples(index=False):
        sheet.append(row)

def count_header_rows(file_path, marker):
    '''
    Description:
        - This function counts the number of rows to skip in a file.
    Parameters:
        - file_path: the path to the file
        - marker: the string to look for in the file
    Returns:
        - The number of rows to skip.
    '''
    with open(file_path, 'r') as f:
        for i, line in enumerate(f):
            if marker in line:
                return i + 1
    return None

def create_chart(data_sheet_name, tmp_df, summary_sheet_name):
    '''
    Description:
        - This function creates a chart on the summary sheet.
    Parameters:
        - data_sheet_name: the name of the sheet containing the data
        - tmp_df: the dataframe containing the data
        - summary_sheet_name: the name of the summary sheet
    Returns:
        - None
    '''
    try:
        chart = LineChart()
        chart.title = "ug/m3 Over Time"
        chart.y_axis.title = "ug/m3"
        chart.x_axis.title = "Time"
        y = Reference(data_sheet_name, min_col=2, min_row=2, max_col=2, max_row=len(tmp_df) + 1)
        x = Reference(data_sheet_name, min_col=7, min_row=2, max_row=len(tmp_df) + 1)
        chart.add_data(y, titles_from_data=True)
        chart.set_categories(x)
        summary_sheet_name.add_chart(chart, "A10")
    except Exception as e:
        print(e)

def create_sheet_with_headers(workbook_name, title, headers):
    '''
    Description:
        - This function creates a sheet with headers.
    Parameters:
        - workbook_name: the name of the workbook
        - title: the title of the sheet
        - headers: the headers for the sheet
    Returns:
        - The sheet.
    '''
    sheet = workbook_name.create_sheet(title=title)
    sheet.append(headers)
    return sheet

def create_summary_sheet(workbook_name, title, summary_stats_var_name):
    '''
    Description:
        - This function creates a sheet with summary statistics.
    Parameters:
        - workbook_name: the name of the workbook
        - title: the title of the sheet
        - summary_stats_var_name: the summary statistics
    Returns:
        - The sheet.
    '''
    sheet = workbook_name.create_sheet(title=title)
    summary_stats_list = summary_stats_var_name.reset_index().values.tolist()
    for row in summary_stats_list:
        sheet.append(row)
    return sheet

def find_directory(drives, target):
    '''
    Description:
        - Function for finding the directory of a particular target.

    Parameters:
        - drives (list) - The drive(s) formatted as 'Letter:\\' (Example: 'C:\\')
        - target (str) - The name of the target folder 

    Returns:
        - Full directory path if it exists, or none if it doesn't
    '''
    for drive in drives:
        directory = os.path.join(drive, target)
        if os.path.exists(directory):
            return directory
    return None

drives = ['%s:\\' % d for d in string.ascii_uppercase if os.path.exists('%s:\\' % d)]
targets = ['CHM\\ExposureScienceLab\\Lead Dust (rlde,hud)', 'ExposureScienceLab\\Lead Dust (rlde,hud)', 'Lead Dust (rlde,hud)']

for target in targets:
    directory = find_directory(drives, target)
    if directory is None:
        print(f"Directory '{target}' not found.")
    else:
        print(f"Directory found: {directory}")
        break

DEFAULT_RLDE_INPUT_DIRECTORY = os.path.join(directory, "Data\\pDR\\new_files")
DEFAULT_RLDE_OUTPUT_DIRECTORY = os.path.join(directory, "Data\\pDR")
CUSTOM_HEADERS = ["record", "ug/m3", "Temp", "RHumidity", "AtmoPressure", "Flags", "time", "date"]
PDR_NUM_DICT = {
    "0115250158": "pdr_1",
    "0115249628": "pdr_2",
    "0115249629": "pdr_3",
    "0115250156": "pdr_4",
    # pdr 5
    "CM19342019": "pdr_6",
    # pdr 7
    "CM21092015": "pdr_8",
    "CM21102014": "pdr_9",
}

def pdr_summary_stats(input_directory, output_directory, *args, **kwargs):
    rlde_flag = kwargs.get('rlde_flag', None)
    
    for file in os.listdir(input_directory):
    
        if file.endswith(".txt"):
            
            # first create some file paths and extract some info
            file_path = os.path.join(input_directory, file) # create file path
            print(f"File found at: {file_path}.")
            base_name = os.path.splitext(os.path.basename(file))[0] # extracting name for xlsx file
            print(f"File name is: {base_name}.")

            if rlde_flag:
                msg = f"Analyzing for RLDE: {rlde_flag}"
                print(msg)  
                subject_match = re.search(r"Subject \d+", base_name) # extracting subject ID. should be in format "Subject #"
    
                if subject_match: # check to make sure we got a subject id from the file name
                    subject_id = subject_match.group(0)
                    print(f"Subject ID found: {subject_id}.")
                    
                    # create a folder for the subject ID if it doesn't exist
                    output_folder = os.path.join(output_directory, subject_id)
                    print(f"This is the output folder: {output_folder}")
                    os.makedirs(output_folder, exist_ok=True)

                    # copy file to correct subject folder
                    copy = shutil.copy2(file_path, output_folder)
                    setwd = os.chdir(output_folder)

                    # create output file paths
                    output_txt_path = os.path.join(output_folder, file)
                    output_xlsx_path = os.path.join(output_folder, f"{base_name}.xlsx")
                else:
                    print(f"Unable to find subject ID for {file_path}.")
                    print(f"File was read as \"{base_name}.\"")
                    print("Please ensure that the file name contains the subject ID as 'Subject #', where # is the subject number.")
                    break
            
            else:
                # copy file to folder
                shutil.copy2(file_path, output_directory)

                # create output file paths
                output_txt_path = os.path.join(output_directory, file)
                output_xlsx_path = os.path.join(output_directory, f"{base_name}.xlsx")

            # working w the xlsx file
            workbook = Workbook()

            # open the file and see if there is a serial number, if so, then go ahead with data extraction
            with open(output_txt_path, 'r') as z:
                lines = z.readlines()
                for line in lines: # parse file for serial number by looking for string after "Serial no.  "
                    if "Serial no." in line: # if serial number is found, then go ahead with data extraction
                        info_row_count = count_header_rows(output_txt_path, 'record')
                        print(f"Number of rows to skip: {info_row_count}.")
                        serial_num = line.split("Serial no.  \", ")[1].strip().replace('"', '')
                        tmp = pd.read_csv(file_path, sep=",", skiprows=info_row_count, header=None, names=CUSTOM_HEADERS)
                        tmp["pdr name"] = PDR_NUM_DICT[serial_num]
                        
                        # remove leading and trailing whitespace from all columns
                        for col in tmp.columns:  
                            if tmp[col].dtype == object:
                                tmp[col] = tmp[col].str.strip()

                        # convert to numeric then calculate summary statistics
                        tmp["ug/m3"] = pd.to_numeric(tmp["ug/m3"], errors='coerce')
                        summary_stats = tmp["ug/m3"].describe()

                        # # remove outliers using IQR method and save removed outliers to new dataframe
                        # q1 = tmp["ug/m3"].quantile(0.25)
                        # q3 = tmp["ug/m3"].quantile(0.75)
                        # iqr = q3 - q1
                        # outliers_iqr_df = tmp[((tmp["ug/m3"] < (q1 - 1.5 * iqr)) | (tmp["ug/m3"] > (q3 + 1.5 * iqr)))]
                        # tmp_no_outliers_iqr = tmp[~((tmp["ug/m3"] < (q1 - 1.5 * iqr)) | (tmp["ug/m3"] > (q3 + 1.5 * iqr)))]
                        # summary_stats_no_outliers_iqr = tmp_no_outliers_iqr["ug/m3"].describe()

                        # # remove outliers using z-score method
                        # outliers_zscore_df = tmp[(tmp["ug/m3"] - tmp["ug/m3"].mean()).abs() > (3 * tmp["ug/m3"].std())]
                        # tmp_no_outliers_zscore = tmp[(tmp["ug/m3"] - tmp["ug/m3"].mean()).abs() <= (3 * tmp["ug/m3"].std())]
                        # summary_stats_no_outliers_zscore = tmp_no_outliers_zscore["ug/m3"].describe()
                        
                        # write the data to xlsx file
                        data_sheet = create_sheet_with_headers(workbook, 'Data_raw', CUSTOM_HEADERS)
                        # data_sheet_no_outliers_iqr = create_sheet_with_headers(workbook, 'Data_no_outliers_iqr', CUSTOM_HEADERS)
                        # data_sheet_no_outliers_zscore = create_sheet_with_headers(workbook, 'Data_no_outliers_zscore', CUSTOM_HEADERS)
                        # outliers_sheet_iqr = create_sheet_with_headers(workbook, 'Outliers_iqr', CUSTOM_HEADERS)
                        # outliers_sheet_zscore = create_sheet_with_headers(workbook, 'Outliers_zscore', CUSTOM_HEADERS)

                        append_rows_to_sheet(data_sheet, tmp)
                        # append_rows_to_sheet(data_sheet_no_outliers_iqr, tmp_no_outliers_iqr)
                        # append_rows_to_sheet(data_sheet_no_outliers_zscore, tmp_no_outliers_zscore)
                        # append_rows_to_sheet(outliers_sheet_iqr, outliers_iqr_df)
                        # append_rows_to_sheet(outliers_sheet_zscore, outliers_zscore_df)

                        summary_sheet = create_summary_sheet(workbook, 'Summary_statistics_raw', summary_stats)
                        # summary_sheet_no_outliers_iqr = create_summary_sheet(workbook, 'Sum_stats_no_outliers_iqr', summary_stats_no_outliers_iqr)
                        # summary_sheet_no_outliers_zscore = create_summary_sheet(workbook, 'Sum_stats_no_outliers_zscore', summary_stats_no_outliers_zscore)
                        
                        # create the chart on summary sheet
                        create_chart(data_sheet, tmp, summary_sheet)
                        # create_chart(data_sheet_no_outliers_iqr, tmp_no_outliers_iqr, summary_sheet_no_outliers_iqr)
                        # create_chart(data_sheet_no_outliers_zscore, tmp_no_outliers_zscore, summary_sheet_no_outliers_zscore)

                        # remove the default sheet
                        default_sheet = workbook["Sheet"]
                        workbook.remove(default_sheet)

                        workbook.save(output_xlsx_path)
                        print(f"Data and Summary statistics saved to {output_xlsx_path}")
        
        if file in os.listdir(output_directory):
            if os.path.getsize(output_xlsx_path) > 0:
                os.remove(file_path)
                print(f"File \"{file}\" deleted from \"{input_directory}.\"")
            else:
                print(f"File \"{file}\" is empty. Please check the file and try again.")

def main():
    parser = argparse.ArgumentParser(description="Extract data from pDR txt files, calculate summary statistics, and create excel file with data, summary stats, and chart.")
    script_dir = os.path.dirname(os.path.realpath(__file__))
    default_input_directory = os.path.join(script_dir, "input")
    default_output_directory = os.path.join(script_dir, "output")
    print(f"script dir: {script_dir}")
    print(f'Default input dir: {default_input_directory}\nDefault output dir: {default_output_directory}')
    parser.add_argument("-i", "--input_directory", type=str, default=default_input_directory, help="The directory containing the pDR txt files.")
    parser.add_argument("-o", "--output_directory", type=str, default=default_output_directory, help="The directory to save the output files to.")
    parser.add_argument("-r", "--rlde", action='store_true', help="Option for running on new RLDE data.")
    
    args = parser.parse_args()
    
    if args.rlde:
        # if the input/output dirs are still default, change it to the rlde ones
        args.input_directory = DEFAULT_RLDE_INPUT_DIRECTORY if args.input_directory == default_input_directory else args.input_directory 
        args.output_directory = DEFAULT_RLDE_OUTPUT_DIRECTORY if args.output_directory == default_output_directory else args.output_directory
        pdr_summary_stats_rlde(args.input_directory, args.output_directory)
    else:
        pdr_summary_stats(args.input_directory, args.output_directory)

if __name__ == "__main__":
    main()