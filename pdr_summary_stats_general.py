#!/usr/bin/env python3

#################################################################################################
#################################################################################################
# Michigan State University - Exposure Science Lab
# Creator: Tuan Nguyen
# Date Created: 2023-10-04
#
# Description:
#  - This is a script for extracting data from the pDR txt files, calculating summary statistics, 
#    and then creating an excel (.xlsx) file with the data, summary stats, and chart.
#  - The script will also remove outliers using the IQR method and the z-score method, and 
#    create a new sheet with the data, summary stats, and chart for each method.
#
# Requirements:
#  - Data should be in the default pDR output format (comma separated, and in a .txt).
#  - pDRs must be in the pdr_num_dict dictionary, so if a new pDR is used, add the serial
#    number to the dictionary.
#
# Usage:
#  - Open the command line and navigate to the directory containing the script.
#  - Here is the command to run the script:
#  - python path\to\pdr_summary_stats_general.py -i path\to\input_directory -o path\to\output_directory
#  - "python" allows us to run the script in the command line.
#  - "path\to\pdr_summary_stats_general.py" is the path to the script.
#  - "-i" is the flag for the input directory.
#  - "path\to\input_directory" is the path to directory containing all your files.
#  - "-o" is the flag for the output directory.
#  - "path\to\output_directory" is the path to where you want the files to be output.
# Example:
#  - python S:\ExposureScienceLab\scripts\pdr_summary_stats_general.py -i "S:\ExposureScienceLab\Lead Dust\Data\pDR\new_files" -o "S:\ExposureScienceLab\Lead Dust\Data\pDR"
#
# Notes:
#  - By default, the input directory is "S:\ExposureScienceLab\scripts\input" and the output directory is "S:\ExposureScienceLab\scripts\output".
#################################################################################################
#################################################################################################

import pandas as pd
import os
import argparse
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference 
import shutil

# variables
custom_headers = ["record", "ug/m3", "Temp", "RHumidity", "AtmoPressure", "Flags", "time", "date"]
pdr_num_dict = {
    "0115250158": "pdr_1",
    "0115249628": "pdr_2",
    "0115249629": "pdr_3",
    "0115250156": "pdr_4",
    # pdr 5
    "CM19342019": "pdr_6",
    # pdr 7
    "CM21092015": "pdr_8",
    # pdr 9
}

# functions
def append_rows_to_sheet(sheet, df):
    for row in df.itertuples(index=False):
        sheet.append(row)

def create_chart(data_sheet_name, tmp_df, summary_sheet_name):
    try:
        chart = LineChart()
        chart.title = "ug/m3 Over Time"
        chart.y_axis.title = "ug/m3"
        chart.x_axis.title = "Time"
        y = Reference(data_sheet_name, min_col=2, min_row=2, max_col=2, max_row=len(tmp_df) + 1)
        x = Reference(data_sheet_name, min_col=7, min_row=2, max_row=len(tmp_df) + 1)
        chart.add_data(y, titles_from_data=True)
        chart.set_categories(x)
        summary_sheet_name.add_chart(chart, "A10")
    except Exception as e:
        print(e)

def create_sheet_with_headers(workbook_name, title, headers):
    sheet = workbook_name.create_sheet(title=title)
    sheet.append(headers)
    return sheet

def create_summary_sheet(workbook_name, title, summary_stats_var_name):
    sheet = workbook_name.create_sheet(title=title)
    summary_stats_list = summary_stats_var_name.reset_index().values.tolist()
    for row in summary_stats_list:
        sheet.append(row)
    return sheet

def pdr_summary_stats(input_directory, output_directory):
    for file in os.listdir(input_directory):
        
        if file.endswith(".txt"):
            
            # first create some file paths and extract some info
            file_path = os.path.join(input_directory, file) # create file path
            print(f"File found at: {file_path}.")
            base_name = os.path.splitext(os.path.basename(file))[0] # extracting name for xlsx file
            print(f"File name is: {base_name}.")

            # copy file to folder
            shutil.copy2(file_path, output_directory)

            # create output file paths
            output_txt_path = os.path.join(output_directory, file)
            output_xlsx_path = os.path.join(output_directory, f"{base_name}.xlsx")

            # working w the xlsx file
            workbook = Workbook()

            # open the file and see if there is a serial number, if so, then go ahead with data extraction
            with open(output_txt_path, 'r') as z:
                lines = z.readlines()
                for line in lines: # parse file for serial number by looking for string after "Serial no.  "
                    if "Serial no." in line: # if serial number is found, then go ahead with data extraction
                        serial_num = line.split("Serial no.  \", ")[1].strip().replace('"', '')
                        tmp = pd.read_csv(file_path, sep=",", skiprows=24, header=None, names=custom_headers)
                        tmp["pdr name"] = pdr_num_dict[serial_num]  # add column for pdr name
                        tmp["time"] = tmp["time"].str.strip()  # Remove leading and trailing whitespace
                        tmp["date"] = tmp["date"].str.strip()
                        summary_stats = tmp["ug/m3"].describe() # summary statistics

                        # remove outliers using IQR method and save removed outliers to new dataframe
                        q1 = tmp["ug/m3"].quantile(0.25)
                        q3 = tmp["ug/m3"].quantile(0.75)
                        iqr = q3 - q1
                        outliers_iqr_df = tmp[((tmp["ug/m3"] < (q1 - 1.5 * iqr)) | (tmp["ug/m3"] > (q3 + 1.5 * iqr)))]
                        tmp_no_outliers_iqr = tmp[~((tmp["ug/m3"] < (q1 - 1.5 * iqr)) | (tmp["ug/m3"] > (q3 + 1.5 * iqr)))]
                        summary_stats_no_outliers_iqr = tmp_no_outliers_iqr["ug/m3"].describe()

                        # remove outliers using z-score method
                        outliers_zscore_df = tmp[(tmp["ug/m3"] - tmp["ug/m3"].mean()).abs() > (3 * tmp["ug/m3"].std())]
                        tmp_no_outliers_zscore = tmp[(tmp["ug/m3"] - tmp["ug/m3"].mean()).abs() <= (3 * tmp["ug/m3"].std())]
                        summary_stats_no_outliers_zscore = tmp_no_outliers_zscore["ug/m3"].describe()
                        
                        # write the data to xlsx file
                        data_sheet = create_sheet_with_headers(workbook, 'Data_raw', custom_headers)
                        data_sheet_no_outliers_iqr = create_sheet_with_headers(workbook, 'Data_no_outliers_iqr', custom_headers)
                        data_sheet_no_outliers_zscore = create_sheet_with_headers(workbook, 'Data_no_outliers_zscore', custom_headers)
                        outliers_sheet_iqr = create_sheet_with_headers(workbook, 'Outliers_iqr', custom_headers)
                        outliers_sheet_zscore = create_sheet_with_headers(workbook, 'Outliers_zscore', custom_headers)

                        append_rows_to_sheet(data_sheet, tmp)
                        append_rows_to_sheet(data_sheet_no_outliers_iqr, tmp_no_outliers_iqr)
                        append_rows_to_sheet(data_sheet_no_outliers_zscore, tmp_no_outliers_zscore)
                        append_rows_to_sheet(outliers_sheet_iqr, outliers_iqr_df)
                        append_rows_to_sheet(outliers_sheet_zscore, outliers_zscore_df)

                        summary_sheet = create_summary_sheet(workbook, 'Summary_statistics_raw', summary_stats)
                        summary_sheet_no_outliers_iqr = create_summary_sheet(workbook, 'Sum_stats_no_outliers_iqr', summary_stats_no_outliers_iqr)
                        summary_sheet_no_outliers_zscore = create_summary_sheet(workbook, 'Sum_stats_no_outliers_zscore', summary_stats_no_outliers_zscore)
                        
                        # create the chart on summary sheet
                        create_chart(data_sheet, tmp, summary_sheet)
                        create_chart(data_sheet_no_outliers_iqr, tmp_no_outliers_iqr, summary_sheet_no_outliers_iqr)
                        create_chart(data_sheet_no_outliers_zscore, tmp_no_outliers_zscore, summary_sheet_no_outliers_zscore)

                        # remove the default sheet
                        default_sheet = workbook["Sheet"]
                        workbook.remove(default_sheet)

                        workbook.save(output_xlsx_path)
                        print(f"Data and Summary statistics saved to {output_xlsx_path}")
        
        if file in os.listdir(output_directory):
            if os.path.getsize(output_xlsx_path) > 0:
                os.remove(file_path)
                print(f"File \"{file}\" deleted from \"{input_directory}.\"")
            else:
                print(f"File \"{file}\" is empty. Please check the file and try again.")

def main():
    parser = argparse.ArgumentParser(description="Extract data from pDR txt files, calculate summary statistics, and create excel file with data, summary stats, and chart.")
    default_input_directory = os.path.expanduser("~/esl_scripts/input")
    default_output_directory = os.path.expanduser("~/esl_scripts/output")
    # default_input_directory = "S:\ExposureScienceLab\scripts\input"
    # default_output_directory = "S:\ExposureScienceLab\scripts\output"
    parser.add_argument("-i", "--input_directory", type=str, default=default_input_directory, help="The directory containing the pDR txt files.")
    parser.add_argument("-o", "--output_directory", type=str, default=default_output_directory, help="The directory to save the output files to.")
    # parser.add_argument("-nr", "--new_rlde_data", type=str, help="Option for running on new RLDE data.")
    args = parser.parse_args()
    pdr_summary_stats(args.input_directory, args.output_directory)

if __name__ == "__main__":
    main()